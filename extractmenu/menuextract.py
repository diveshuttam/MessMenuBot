#edit the argument file for the extraction only one file at a time should be in this directory
import sys
import openpyxl
import json
import datetime
import os
import glob

file_path="../menu/a.json"
def ensure_dir(filepath):
    directory = os.path.dirname(filepath)
    if not os.path.exists(directory):
        os.makedirs(directory)

ensure_dir(file_path)

for wb in glob.glob("*.xlsx"):
    messywb = openpyxl.load_workbook(wb)
    messysheet = messywb.get_active_sheet()
    nd=0
    with open("./arg.json") as json_data:       #this contains arguments
        argu=json.load(json_data)
        json_data.close()
        breakfast_start = bs = int(argu["breakfast_start"])
        breakfast_end = be = int(argu["breakfast_end"])
        lunch_start = ls = int(argu["lunch_start"])
        lunch_end = le = int(argu["lunch_end"])
        dinner_start = ds = int(argu["dinner_start"])
        dinner_end = de = int(argu["dinner_end"])
        no_of_days = nd = int(argu["no_of_days"])

    print(bs,be,ls,le,ds,de,nd)
    for i in range(1, nd+1):
        breakfast = []
        lunch = []
        dinner = []
        dt = messysheet.cell(row=1, column=i).value
        dt.replace(year=dt.today().year)
        filename=dt.strftime("../menu/%d-%m-%Y.json")
        for j in range(bs,be+1):
            # add to breakfast if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                breakfast.append(temp.strip())
        for j in range(ls, le+1):
            # add to breakfast if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                lunch.append(temp.strip())
        for j in range(ds, de+1):
            # add to breakfast if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp != None and temp[0].isalpha()):
                dinner.append(temp.strip())
        menu={} #map menu
        menu["breakfast"]=breakfast
        menu["lunch"]=lunch
        menu["dinner"]=dinner
        json_string=json.dumps(menu,indent=2)
        #print(menu)
        fi=open(filename,"w+")
        fi.write(json_string)
        fi.close()
